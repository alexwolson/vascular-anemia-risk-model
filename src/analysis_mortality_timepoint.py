"""
Task 5: Mortality Timepoint Analysis
=====================================
Inspects DEAD and PROC_SURVIVALDAYS from the merged VQI dataset,
cross-references data dictionaries, and produces summary statistics
and a narrative report characterising the mortality endpoint.
"""

import os
import pathlib
import textwrap

import numpy as np
import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
PARQUET_PATH = DATA_DIR / "processed" / "merged_vqi_2012_2020.parquet"
OUTPUT_DIR = PROJECT_ROOT / "output" / "task5_mortality"

DICT_FILES = [
    DATA_DIR / "INFRA_Data_Dictionary.xlsx",
    DATA_DIR / "OPEN_AAA_Data_Dictionary.xlsx",
    DATA_DIR / "SUPRA_Data_Dictionary.xlsx",
]
MERGED_DESC = DATA_DIR / "merged_descriptor.xlsx"


# ---------------------------------------------------------------------------
# 1. Load data
# ---------------------------------------------------------------------------
print("=" * 72)
print("TASK 5 -- Mortality Timepoint Analysis")
print("=" * 72)

df = pd.read_parquet(PARQUET_PATH)
print(f"\nLoaded {PARQUET_PATH.name}: {df.shape[0]:,} rows x {df.shape[1]} columns")

# ---------------------------------------------------------------------------
# 2. Create output directory
# ---------------------------------------------------------------------------
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
print(f"Output directory: {OUTPUT_DIR}")

# ---------------------------------------------------------------------------
# 3. Inspect DEAD and PROC_SURVIVALDAYS
# ---------------------------------------------------------------------------
print("\n" + "-" * 72)
print("Variable inspection")
print("-" * 72)

print(f"\nDEAD  dtype: {df['DEAD'].dtype}")
print(f"DEAD  value counts (incl. NaN):\n{df['DEAD'].value_counts(dropna=False).to_string()}")

print(f"\nPROC_SURVIVALDAYS  dtype: {df['PROC_SURVIVALDAYS'].dtype}")
print(f"PROC_SURVIVALDAYS  describe:\n{df['PROC_SURVIVALDAYS'].describe().to_string()}")

# Flag negative survival days (data-quality note)
n_neg = (df["PROC_SURVIVALDAYS"] < 0).sum()
if n_neg > 0:
    print(f"\n** Data-quality note: {n_neg} records have negative PROC_SURVIVALDAYS "
          "(likely date-entry errors). These are included in overall stats but "
          "excluded from follow-up-time bins and mortality-window calculations.")

# ---------------------------------------------------------------------------
# 4. Cross-reference data dictionaries
# ---------------------------------------------------------------------------
print("\n" + "-" * 72)
print("Data-dictionary definitions")
print("-" * 72)

dict_definitions: dict[str, list[dict]] = {"DEAD": [], "PROC_SURVIVALDAYS": []}

for fpath in DICT_FILES + [MERGED_DESC]:
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        var_name = str(row[0]).strip() if row[0] else ""
        if var_name in dict_definitions:
            entry = {h: v for h, v in zip(header, row) if v is not None}
            entry["_source_file"] = fpath.name
            dict_definitions[var_name].append(entry)
    wb.close()

for var, entries in dict_definitions.items():
    print(f"\n  {var}:")
    for e in entries:
        src = e.pop("_source_file")
        options = e.get("Select Options", "")
        helptext = e.get("Supplementary Help Text", "")
        print(f"    [{src}]")
        if options:
            print(f"      Options : {options}")
        if helptext:
            wrapped = textwrap.fill(str(helptext), width=70,
                                     initial_indent="      Help    : ",
                                     subsequent_indent="                ")
            print(wrapped)

# ---------------------------------------------------------------------------
# 5. Compute statistics
# ---------------------------------------------------------------------------
print("\n" + "-" * 72)
print("Summary statistics")
print("-" * 72)

total_n = len(df)
dead_series = df["DEAD"]
surv = df["PROC_SURVIVALDAYS"]

# -- Overall mortality -------------------------------------------------------
n_dead = (dead_series == "1").sum()
n_alive = (dead_series == "0").sum()
n_missing = dead_series.isna().sum()
mortality_rate = n_dead / (n_dead + n_alive) * 100

print(f"\nTotal patients           : {total_n:,}")
print(f"  Died  (DEAD=1)         : {n_dead:,}  ({mortality_rate:.2f}%)")
print(f"  Alive (DEAD=0)         : {n_alive:,}  ({100 - mortality_rate:.2f}%)")
print(f"  Missing DEAD           : {n_missing:,}")


def describe_days(series: pd.Series, label: str) -> dict:
    """Return a dict of summary stats for a numeric series."""
    stats = {
        "Group": label,
        "N": int(series.count()),
        "Mean": round(series.mean(), 1),
        "Median": round(series.median(), 1),
        "SD": round(series.std(), 1),
        "Min": int(series.min()),
        "Max": int(series.max()),
    }
    return stats


stats_all = describe_days(surv, "All patients")
stats_dead = describe_days(surv[dead_series == "1"], "Died (DEAD=1)")
stats_alive = describe_days(surv[dead_series == "0"], "Survived (DEAD=0)")

for s in (stats_all, stats_dead, stats_alive):
    print(f"\n  PROC_SURVIVALDAYS -- {s['Group']}:")
    print(f"    N={s['N']:,}  Mean={s['Mean']:.1f}  Median={s['Median']:.1f}  "
          f"SD={s['SD']:.1f}  Min={s['Min']}  Max={s['Max']}")

# -- Follow-up time distribution (exclude negative days) --------------------
surv_pos = surv[surv >= 0]
n_pos = len(surv_pos)

bins = [
    ("< 30 days", surv_pos < 30),
    ("30-365 days", (surv_pos >= 30) & (surv_pos <= 365)),
    ("1-2 years", (surv_pos > 365) & (surv_pos <= 730)),
    ("2-5 years", (surv_pos > 730) & (surv_pos <= 1825)),
    ("5+ years", surv_pos > 1825),
]

print(f"\n  Follow-up time distribution (N={n_pos:,} with PROC_SURVIVALDAYS >= 0):")
followup_rows = []
for label, mask in bins:
    count = int(mask.sum())
    pct = count / n_pos * 100
    print(f"    {label:15s}  n={count:>6,}  ({pct:5.1f}%)")
    followup_rows.append({"Follow-up window": label, "N": count, "%": round(pct, 2)})

# -- 30-day mortality --------------------------------------------------------
mask_30d = (dead_series == "1") & (surv >= 0) & (surv <= 30)
n_30d_dead = int(mask_30d.sum())
rate_30d = n_30d_dead / total_n * 100
print(f"\n  30-day mortality       : {n_30d_dead:,}  ({rate_30d:.2f}%)")

# -- 1-year mortality --------------------------------------------------------
mask_1y = (dead_series == "1") & (surv >= 0) & (surv <= 365)
n_1y_dead = int(mask_1y.sum())
rate_1y = n_1y_dead / total_n * 100
print(f"  1-year mortality       : {n_1y_dead:,}  ({rate_1y:.2f}%)")

# ---------------------------------------------------------------------------
# 6. Save summary CSV
# ---------------------------------------------------------------------------
csv_path = OUTPUT_DIR / "mortality_timepoint_summary.csv"

summary_rows = [
    {"Metric": "Total patients", "Value": total_n},
    {"Metric": "Deaths (DEAD=1)", "Value": n_dead},
    {"Metric": "Alive (DEAD=0)", "Value": n_alive},
    {"Metric": "Missing DEAD", "Value": n_missing},
    {"Metric": "Overall mortality rate (%)", "Value": round(mortality_rate, 2)},
    {"Metric": "", "Value": ""},
    {"Metric": "--- PROC_SURVIVALDAYS: All patients ---", "Value": ""},
    {"Metric": "N", "Value": stats_all["N"]},
    {"Metric": "Mean", "Value": stats_all["Mean"]},
    {"Metric": "Median", "Value": stats_all["Median"]},
    {"Metric": "SD", "Value": stats_all["SD"]},
    {"Metric": "Min", "Value": stats_all["Min"]},
    {"Metric": "Max", "Value": stats_all["Max"]},
    {"Metric": "", "Value": ""},
    {"Metric": "--- PROC_SURVIVALDAYS: Died (DEAD=1) ---", "Value": ""},
    {"Metric": "N", "Value": stats_dead["N"]},
    {"Metric": "Mean", "Value": stats_dead["Mean"]},
    {"Metric": "Median", "Value": stats_dead["Median"]},
    {"Metric": "SD", "Value": stats_dead["SD"]},
    {"Metric": "Min", "Value": stats_dead["Min"]},
    {"Metric": "Max", "Value": stats_dead["Max"]},
    {"Metric": "", "Value": ""},
    {"Metric": "--- PROC_SURVIVALDAYS: Survived (DEAD=0) ---", "Value": ""},
    {"Metric": "N", "Value": stats_alive["N"]},
    {"Metric": "Mean", "Value": stats_alive["Mean"]},
    {"Metric": "Median", "Value": stats_alive["Median"]},
    {"Metric": "SD", "Value": stats_alive["SD"]},
    {"Metric": "Min", "Value": stats_alive["Min"]},
    {"Metric": "Max", "Value": stats_alive["Max"]},
    {"Metric": "", "Value": ""},
    {"Metric": "--- Follow-up time distribution ---", "Value": ""},
]
for fr in followup_rows:
    summary_rows.append({"Metric": fr["Follow-up window"], "Value": f"n={fr['N']}, {fr['%']}%"})
summary_rows.extend([
    {"Metric": "", "Value": ""},
    {"Metric": "30-day mortality count", "Value": n_30d_dead},
    {"Metric": "30-day mortality rate (%)", "Value": round(rate_30d, 2)},
    {"Metric": "1-year mortality count", "Value": n_1y_dead},
    {"Metric": "1-year mortality rate (%)", "Value": round(rate_1y, 2)},
    {"Metric": "Negative PROC_SURVIVALDAYS records", "Value": n_neg},
])

pd.DataFrame(summary_rows).to_csv(csv_path, index=False)
print(f"\nSaved summary CSV -> {csv_path}")

# ---------------------------------------------------------------------------
# 7. Write narrative
# ---------------------------------------------------------------------------
narrative_path = OUTPUT_DIR / "mortality_timepoint_narrative.txt"

# Compute additional stats for the narrative
pct_lt30 = followup_rows[0]["%"]
pct_lt1y = followup_rows[0]["%"] + followup_rows[1]["%"]
pct_gt1y = 100 - pct_lt1y
median_surv_dead = surv[dead_series == "1"].median()
median_surv_alive = surv[dead_series == "0"].median()

narrative = textwrap.dedent(f"""\
    Mortality Timepoint Narrative -- Task 5
    ========================================

    Data source : merged_vqi_2012_2020.parquet
    Records     : {total_n:,}
    Date range  : VQI procedures 2012-2020

    1. What the DEAD variable captures
    -----------------------------------
    According to every VQI data dictionary consulted (INFRA, OPEN-AAA, SUPRA,
    and the merged descriptor), the DEAD variable records the "patient's last
    known mortality status" with a default value of 0 (Alive). Its coded values
    are 0 = No (alive at last contact) and 1 = Yes (known dead).

    This is NOT a fixed-timepoint endpoint (it is not strictly 30-day mortality
    or 1-year mortality). Instead, DEAD represents last-known vital status --
    essentially a vital-status flag that is updated whenever new follow-up or
    linkage data become available. Some patients may have been followed for only
    a few days (e.g. lost to follow-up after discharge), while others were
    tracked for over 16 years.

    2. Mortality type classification
    ---------------------------------
    DEAD is best characterised as "last-known-status mortality" (sometimes
    called all-cause mortality at last follow-up). It should be interpreted in
    conjunction with PROC_SURVIVALDAYS, which provides the time horizon for
    each patient.

    The data dictionary defines PROC_SURVIVALDAYS as:
      "The longest time of survival data available for the patient.
       Survival days are calculated as the Last Date of Contact (or Date of
       Death) minus the Procedure date."

    This means:
      - For patients who died (DEAD=1): PROC_SURVIVALDAYS approximates the
        time from procedure to death.
      - For patients alive (DEAD=0): PROC_SURVIVALDAYS approximates the time
        from procedure to last known contact, and the true survival time is
        right-censored at that point.

    Because follow-up length is variable, the raw DEAD rate of {mortality_rate:.1f}%
    ({n_dead:,}/{n_dead + n_alive:,}) cannot be compared directly to a
    standardised 30-day or 1-year mortality rate.

    Crude timepoint mortality estimates (not survival-analysis adjusted):
      - 30-day mortality : {n_30d_dead:,} deaths / {total_n:,} patients = {rate_30d:.2f}%
      - 1-year mortality : {n_1y_dead:,} deaths / {total_n:,} patients = {rate_1y:.2f}%

    These are lower bounds because patients censored before those time points
    might have died after censoring.

    3. Follow-up period characteristics
    ------------------------------------
    Overall PROC_SURVIVALDAYS (all {total_n:,} patients):
      Mean   = {stats_all['Mean']:.1f} days  ({stats_all['Mean']/365.25:.1f} years)
      Median = {stats_all['Median']:.1f} days  ({stats_all['Median']/365.25:.1f} years)
      SD     = {stats_all['SD']:.1f} days
      Range  = {stats_all['Min']} to {stats_all['Max']} days

    Among patients who died (DEAD=1, n={stats_dead['N']:,}):
      Mean   = {stats_dead['Mean']:.1f} days  ({stats_dead['Mean']/365.25:.1f} years)
      Median = {stats_dead['Median']:.1f} days  ({stats_dead['Median']/365.25:.1f} years)

    Among survivors (DEAD=0, n={stats_alive['N']:,}):
      Mean   = {stats_alive['Mean']:.1f} days  ({stats_alive['Mean']/365.25:.1f} years)
      Median = {stats_alive['Median']:.1f} days  ({stats_alive['Median']/365.25:.1f} years)

    Follow-up time distribution:
      < 30 days   : {followup_rows[0]['N']:>6,} ({followup_rows[0]['%']:5.1f}%)
      30-365 days : {followup_rows[1]['N']:>6,} ({followup_rows[1]['%']:5.1f}%)
      1-2 years   : {followup_rows[2]['N']:>6,} ({followup_rows[2]['%']:5.1f}%)
      2-5 years   : {followup_rows[3]['N']:>6,} ({followup_rows[3]['%']:5.1f}%)
      5+ years    : {followup_rows[4]['N']:>6,} ({followup_rows[4]['%']:5.1f}%)

    Roughly {pct_lt30:.0f}% of patients have less than 30 days of follow-up and
    {pct_gt1y:.0f}% have more than one year. The wide spread in follow-up time
    means that any mortality analysis should account for censoring, ideally via
    Kaplan-Meier estimates or Cox proportional-hazards modelling rather than
    simple proportions.

    Data-quality note: {n_neg} records have negative PROC_SURVIVALDAYS, which
    likely reflect date-entry errors (death or contact date recorded before the
    procedure date). These were excluded from follow-up-time bins and
    time-window mortality calculations but retained in the overall dataset.
""")

narrative_path.write_text(narrative)
print(f"Saved narrative   -> {narrative_path}")

print("\n" + "=" * 72)
print("Task 5 complete.")
print("=" * 72)
